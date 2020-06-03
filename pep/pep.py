import xlsxwriter
import openpyxl
from pathlib import Path
import sys
from PyQt5.QtWidgets import (QWidget, QLabel, QComboBox, QTableWidget, QTableWidgetItem, QApplication, QGridLayout, QAction, QMainWindow, qApp, QPushButton, QDialog, QVBoxLayout, QTextEdit, QHeaderView, QApplication, QWidget, QInputDialog, QLineEdit, QFileDialog
)
from PyQt5.QtGui import QIcon, QPainter, QColor, QPen
from PyQt5.QtCore import pyqtSlot, Qt
import sqlite3
from datetime import timedelta, datetime, tzinfo, timezone
from PyQt5 import QtCore, QtGui, QtWidgets


class Data(): 
    nom="pep.db"
    tabGauche= []
    tabCentrale= []

    def __init__(self):
        self.conn= sqlite3.connect(self.nom)
        self.c= self.conn.cursor()

    def importMilitaire(self, lieu):
        self.tabGauche= []
        self.tabCentrale= []
        res= self.c.execute("SELECT * FROM militaire WHERE lieu LIKE '"+lieu+"' order by id")
        for ligne in res:
            self.tabGauche.append([ligne[1], ligne[2], ligne[3]])
            self.tabGauche.append([" ", " ", ""])
            self.tabCentrale.append(ligne[4].split(","))
            self.tabCentrale.append(ligne[5].split(",")) 
        self.tabGauchePlanningsBlocs=[self.tabGauche, self.tabCentrale]
        return self.tabGauchePlanningsBlocs

    def exportMilitaire(self, ntabGauche, ntabCentrale, lieu):
        jointure= ","
        for i in range(int(len(ntabGauche)/2)):
            self.c.execute("UPDATE militaire SET grade= '"+ntabGauche[i*2][0]+"', nom= '"+ntabGauche[i*2][1]+"', prenom= '"+ntabGauche[i*2][2]+"', planning= '"+jointure.join(ntabCentrale[i*2])+"', bloc= '"+jointure.join(ntabCentrale[(i*2)+1])+"' WHERE (id = "+ str(i)+" and lieu= '"+lieu+"')")
        self.conn.commit()

    def insert(self, chaine, lieu):
        self.c.execute("INSERT INTO militaire (id, grade, nom, prenom, planning, bloc, lieu) VALUES ("+chaine+",'"+lieu+"')")
        self.c.execute("DELETE FROM militaire WHERE rowid NOT IN (SELECT min(rowid) FROM militaire GROUP BY grade, nom, prenom, lieu)")
        self.conn.commit()

    def remove(self, ligne, lieu):
        self.c.execute("DELETE FROM militaire WHERE id= "+str(ligne)+" AND lieu= '"+lieu+"'")
        self.conn.commit()

    def resetID(self, ntabGauche, lieu):
        for i in range(len(ntabGauche)):
            if i%2 == 0:
                self.c.execute("UPDATE militaire SET id="+str(int(i/2))+" WHERE (nom= '"+ntabGauche[i][1]+"' AND prenom= '"+ntabGauche[i][2]+"' AND lieu= '"+lieu+"')")
        self.conn.commit()
    def viderToutLesMilitairesDUnLieu(self, lieu):
        self.c.execute("DELETE FROM militaire WHERE lieu= '"+lieu+"'")
        self.conn.commit()

    def getLieu(self):
        res= self.c.execute("SELECT lieu FROM options")
        for i in res:
            print(i[0])
            res2= i[0]
        return res2

    def saveLieu(self, chaine):
        res= self.c.execute("Update options set lieu= '"+chaine+"'")
        self.conn.commit()

    def ajouterUnMilitaireParDefaut(self, lieu):
        tab= [""]*56
        chaine= "0, 'grade', 'nom', 'prenom', '"+(",".join(tab))+"', '"+(",".join(tab))+"', '"+lieu+"'"
        self.c.execute("INSERT INTO militaire (id, grade, nom, prenom, planning, bloc, lieu) VALUES ("+chaine+")")
        self.c.execute("DELETE FROM militaire WHERE rowid NOT IN (SELECT min(rowid) FROM militaire GROUP BY grade, nom, prenom, lieu)")
        self.conn.commit()

class App(QWidget):

    def __init__(self):
        super().__init__()
        self.tabGauche= []
        self.tabDroite= []
        self.tabLieu= []
        self.title = 'Bat hôp 2 planning'
        self.left = 0
        self.top = 0
        self.width = 1300
        self.height = 700
        self.debut=0
        self.fin= 56
        #les valeurs jours et nuit
        self.minutesNuit= 784
        self.minutesJour= 720
        #les couleurs
        self.rouge= QColor(255, 118, 160)
        self.gris= QColor(178, 190, 195)
        self.gris2= QColor(223, 230, 233)
        self.jaune= QColor(246, 229, 141)
        self.vert= QColor(123, 237, 159)
        self.bleu= QColor(116, 185, 255)
        self.blanc= QColor(255, 255, 255)
        #lieu par Defaut
        self.lieu="Delemont"
        #taille par défaut
        self.cinquanteSixEspace= ",".join([" "]*56)
        #les tables
        self.initUI()

    #fonction d'initialisation
    def initUI(self):
        #définition de la page de base
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)
        #Définition de la couleur de fond
        p = self.palette()
        p.setColor(self.backgroundRole(), QColor(255, 255, 255))
        self.setPalette(p)
        #charger les données dans deux tableaux (gauche et central)
        self.initOptions()
        self.initTable()
        self.initTab()
        self.chargerToutesLesTab()
        self.display()

    #2 fonctions pour la base de donnée
    def loadData(self):
        #self.data= Data()
        self.tabGauchePlanningsBlocs= self.data.importMilitaire(self.lieu)
        self.tabGauche= self.tabGauchePlanningsBlocs[0]
        self.tabCentrale= self.tabGauchePlanningsBlocs[1]

    def saveData(self):
        self.data.exportMilitaire(self.tabGauche, self.tabCentrale, self.lieu)
        
    def chargerToutesLesTab(self):
        self.chargerTable(self.tabCentrale, self.tableCentrale)
        self.chargerTable(self.tabGauche, self.tableGauche)
        self.chargerTable2(self.tabDroite, self.tableDroite)
        self.chargerTable(self.tabBas, self.tableBas)
        self.chargerTable(self.tabHaute, self.tableHaute)
        self.chargerTable(self.tabCoin, self.tableCoin)
        self.marquerTable(self.tableCentrale)
        self.zoneCouleur() 

    def zoneCouleur(self):
        #zone pour couleur
        #on va mettre des bandes grises
        for ligne in range(1, int(len(self.tabCentrale)), 2):
            #bandes grises pour la table centrale
            self.colorier(self.selectLigne(ligne, self.tabCentrale), self.tableCentrale, self.gris2)
            #bandes grises pour la table de gauche
            self.colorier(self.selectLigne(ligne, self.tabGauche), self.tableGauche, self.gris2)
            #bandes grises pour la table de droite
            self.colorier([[ligne, 0]], self.tableDroite, self.gris2)
            #verouillage pour la table centrale
            self.verouiller([[ligne-1, 0]], self.tableDroite)
            #verouillage pour la table de droite
            self.verouiller([[ligne, 0]], self.tableDroite)
            #verouillage pour la table de gauche
            self.verouiller(self.selectLigne(ligne, self.tabGauche), self.tableGauche)

        #verouillage pour la table du haut
        for ligne in range(2):
            self.verouiller(self.selectLigne(ligne, self.tabHaute), self.tableHaute)

        #verouillage pour la table du haut
        for ligne in range(5):
            self.verouiller(self.selectLigne(ligne, self.tabBas), self.tableBas)

        #coloration des dimanches
        for colonne in range(6, 56, 7):
            self.colorier([[1, colonne]], self.tableHaute, self.rouge)

        #coloration du tabcoin
        self.colorier([[0, 2]], self.tableCoin, self.bleu)
        self.colorier([[2, 2]], self.tableCoin, self.jaune)
        self.colorier([[3, 2]], self.tableCoin, self.vert)
        self.colorier([[4, 2]], self.tableCoin, self.gris)

    def zoneCouleur2(self, nTabCentrale, nTabHaute):
        #zone pour couleur
        #on va mettre des bandes grises
        for ligne in range(1, int(len(nTabCentrale)), 2):
            #bandes grises pour la table centrale
            self.colorier(self.selectLigne(ligne, nTabCentrale), self.tableCentrale, self.gris2)
            #bandes grises pour la table de gauche
            self.colorier(self.selectLigne(ligne, self.tabGauche), self.tableGauche, self.gris2)
            #bandes grises pour la table de droite
            self.colorier([[ligne, 0]], self.tableDroite, self.gris2)
            #verouillage pour la table centrale
            self.verouiller([[ligne-1, 0]], self.tableDroite)
            #verouillage pour la table de droite
            self.verouiller([[ligne, 0]], self.tableDroite)
            #verouillage pour la table de gauche
            self.verouiller(self.selectLigne(ligne, self.tabGauche), self.tableGauche)

        #verouillage pour la table du haut
        for ligne in range(2):
            self.verouiller(self.selectLigne(ligne, nTabHaute), self.tableHaute)

        #coloration des dimanches
        for colonne in range(6, len(nTabHaute), 7):
            self.colorier([[1, colonne]], nTabHaute, self.rouge)

        #coloration du tabcoin
        self.colorier([[0, 2]], self.tableCoin, self.bleu)
        self.colorier([[2, 2]], self.tableCoin, self.jaune)
        self.colorier([[3, 2]], self.tableCoin, self.vert)
        self.colorier([[4, 2]], self.tableCoin, self.gris2)

    def chargerTable(self, tab, table):
        #dimensions (nb cellules) de la table
        ligneTab= len(tab)
        colonneTab= len(tab[0])
        ligneTable= table.rowCount()
        colonneTable= table.columnCount()

        #on arrange les lignes
        #si on a moins de ligne que la table
        if ligneTab < ligneTable:
            nbLigneEnMoins= ligneTable-ligneTab
            for i in range(nbLigneEnMoins, 0, -1):
                table.removeRow(table.rowCount()-1)
        elif ligneTab > ligneTable:
            nbLigneEnPlus= ligneTab-ligneTable
            for i in range(0, nbLigneEnPlus, 1):
                table.insertRow(table.rowCount())
        #on arrange les colonnes
        #si on a moins de colonne que la table
        if colonneTab < colonneTable:
            nbColonneEnMoins= colonneTable-colonneTab
            for i in range(nbColonneEnMoins, 0, -1):
                table.removeColumn(table.columnCount()-1)
        elif colonneTab > colonneTable:
            nbColonneEnPlus= colonneTab-colonneTable
            for i in range(0, nbColonneEnPlus, 1):
                table.insertColumn(table.columnCount())
    
        #remplissage de la table
        for i in range(ligneTab):
            for j in range(colonneTab):
                table.setItem(i, j, QTableWidgetItem(str(tab[i][j])))

    def chargerTable2(self, tab, table):
        #dimensions (nb cellules) de la table
        ligneTab= len(tab)
        colonneTab= 1
        ligneTable= table.rowCount()
        colonneTable= table.columnCount()

        #on arrange les lignes
        #si on a moins de ligne que la table
        if ligneTab < ligneTable:
            nbLigneEnMoins= ligneTable-ligneTab
            for i in range(nbLigneEnMoins, 0, -1):
                table.removeRow(table.rowCount()-1)
        elif ligneTab > ligneTable:
            nbLigneEnPlus= ligneTab-ligneTable
            for i in range(0, nbLigneEnPlus, 1):
                table.insertRow(table.rowCount())
        #on arrange les colonnes
        #si on a moins de colonne que la table
        if colonneTab < colonneTable:
            nbColonneEnMoins= colonneTable-colonneTab
            for i in range(nbColonneEnMoins, 0, -1):
                table.removeColumn(table.columnCount()-1)
        elif colonneTab > colonneTable:
            nbColonneEnPlus= colonneTab-colonneTable
            for i in range(0, nbColonneEnPlus, 1):
                table.insertColumn(table.columnCount())
    
        #remplissage de la table
        for i in range(ligneTab):
            for j in range(colonneTab):
                table.setItem(i, j, QTableWidgetItem(str(tab[i])))

    def creerTable(self, largeur, hauteur):
       # creer la table
        tableWidget = QTableWidget()
        #dimensions (nb cellules) de la table
        tableWidget.setMaximumWidth(largeur)
        tableWidget.setMaximumHeight(hauteur)
        tableWidget.setRowCount(1)
        tableWidget.setColumnCount(1)

        #on cache les headers
        tableWidget.verticalHeader().hide()
        tableWidget.horizontalHeader().hide()


        return tableWidget


    def display(self):
        #liste déroulante
        self.combobox1= QComboBox()
        self.combobox1.addItems(self.tabLieu)
        self.combobox1.currentTextChanged.connect(self.comboLieu)

        self.combobox2= QComboBox()
        self.combobox2.addItems(["Tous", "semaine 1", "semaine 2", "semaine 3", "semaine 4", "semaine 5", "semaine 6", "semaine 7", "semaine 8"])
        self.combobox2.currentTextChanged.connect(self.comboSemaine)

        self.layout = QGridLayout()

        #initialise les boutons/fonctions pour l'option
        options= self.initButton("Plus d'options", self.plusDOption) 

        self.layout.addWidget(options, 0, 0) 
        self.layout.addWidget(self.combobox1, 1, 0)

        self.photo = QLabel()
        self.photo.setGeometry(QtCore.QRect(0, 0, 20, 20))
        self.photo.setPixmap(QtGui.QPixmap("pep.jpg"))
        self.photo.setScaledContents(True)
        self.photo.setObjectName("photo")

        self.layoutMain = QGridLayout()

        self.layoutMain.addWidget(self.photo, 2, 2) 
        self.layoutMain.addLayout(self.layout, 0, 0)
        self.layoutMain.addWidget(self.tableHaute, 0, 1) 
        self.layoutMain.addWidget(self.combobox2, 0, 2) 
        self.layoutMain.addWidget(self.tableGauche, 1, 0) 
        self.layoutMain.addWidget(self.tableCentrale, 1, 1) 
        self.layoutMain.addWidget(self.tableDroite, 1, 2) 
        self.layoutMain.addWidget(self.tableCoin, 2, 0) 
        self.layoutMain.addWidget(self.tableBas, 2, 1) 

        #Début
        self.setLayout(self.layoutMain) 

        #création des callback
        self.tableCentrale.itemChanged.connect(self.modifierData)
        self.tableGauche.itemChanged.connect(self.modifierDataTabGauche)
        self.tableHaute.cellDoubleClicked.connect(self.journee)

        # Show widget
        self.show()

    def initTable(self):
        #changer création des tables
        self.tableCentrale= self.creerTable(800, 400)
        self.tableGauche= self.creerTable(300, 400)
        self.tableHaute= self.creerTable(800, 100)
        self.tableDroite= self.creerTable(200, 400)
        self.tableBas= self.creerTable(800, 170)
        self.tableCoin= self.creerTable(300, 170)

        self.synchronisation()


    def initTab(self):
        #tabCentrale et tabGauche sont déjà définies
        #car elles sont importées depuis la base de donnée
        #tabCoin et tabHautes sont définis par défaut
        #tabBas et tabDroite sont calculées
        
        #pour tabGauche et tabCentrale
        self.loadData()
        #pour tableHaute
        self.definirInterval()
        #pour tableDroite
        self.heureTotal()
        #pour tableBas
        self.initTabBas()
        #pour tableCoin
        self.initTabCoin()

    def heureTotal(self):
        self.tabDroite= []
        #table droite
        for ligne in range(len(self.tabCentrale)):
            #self.heureTotal(self.tabGauche[ligne], ligne)
            if ligne%2 == 0:
                chaine= self.tabCentrale[ligne]
                heure= chaine.count('N')*self.minutesNuit+(chaine.count('J')*self.minutesJour)
                self.tabDroite.append(self.minuteAHeureMinute(heure))
            else:
                self.tabDroite.append(" ")

    def heureTotal2(self, table, debut, fin):
        tab2= []
        #table droite
        for ligne in range(len(self.tabCentrale)):
            #self.heureTotal(self.tabGauche[ligne], ligne)
            chaine= self.tabCentrale[ligne][debut:fin]
            heure= chaine.count('N')*self.minutesNuit+(chaine.count('J')*self.minutesJour)
            tab2.append(self.minuteAHeureMinute(heure))
        return tab2

    def splitParCaractere(self, a):
        A=[]
        for k in range(len(a)):
            #a est une chaine de caractère
            A.append([a[k][i:i+1] for i in range(0, len(a))])
        return A

    def clearLayout(self):
        for i in reversed(range(self.layoutMain.count())): 
                self.layoutMain.itemAt(i).widget().setParent(None)

    def initTabBas(self):
       #on definit la longueur de l'interval
       longueurInterval= 56
       #on calcul déjà le contenu de la table
       #on compte le nombre de N, R, J, I et C dans chaque colonne
       self.N= [0]*longueurInterval
       self.R= [0]*longueurInterval
       self.J= [0]*longueurInterval
       self.I= [0]*longueurInterval
       self.C= [0]*longueurInterval
       #on garde ces valeurs dans les variables du même nom (self.N, self.R, etc.)
       for colonne in range(longueurInterval):
            for ligne in range(len(self.tabCentrale)):
                self.compter(self.tabCentrale[ligne][colonne], colonne)
       self.tabBas= [self.N, self.R, self.J, self.I, self.C]

    def initTabCoin(self):
       #on definit la longueur de l'interval
       longueurInterval= 56
       # Create table
       self.tabCoin= []
       self.tabCoin.append(["13:04", "Nuit", "N"])
       self.tabCoin.append(["-", "Repos", "R"])
       self.tabCoin.append(["12:00", "Jour", "J"])
       self.tabCoin.append(["-", "Instruction", "I"])
       self.tabCoin.append(["-", "Compensation", "C"])


    @pyqtSlot()

    def colorier(self, positions, table, couleur):
        for position in positions:
            table.item(position[0], position[1]).setBackground(couleur)

    def verouiller(self, positions, table):
        for position in positions:
            table.item(position[0], position[1]).setFlags(Qt.ItemIsEnabled)

    def selectLigne(self, ligne, table):
        position= []
        for colonne in range(len(table[0])):
            position.append([ligne, colonne])
        return position

    def selectColonne(self, colonne, table):
        position= []
        for ligne in range(len(table)):
            position.append([ligne, colonne])
        return position

    def regles(self, ligne, colonne):
       #on calcul d'abord la liste des totaux en heure
       if ligne%2 == 0:
           self.marquerLigne(self.tableCentrale, ligne)
           self.erreur(ligne, "JJJJ")
           self.erreur(ligne, "NNN")

    def erreur(self, ligne, chaine):
       #on calcul d'abord la liste des totaux en heure
       planning= "".join(self.tabCentrale[ligne])
       i= 0
       colonneOuEstLErreur= [] 
       while i < len(planning) and planning[i:].count(chaine) > 0:
          i= planning.index(chaine, i)
          colonneOuEstLErreur.append(i)
          i= i+4
       self.marquerErreur(ligne, colonneOuEstLErreur, len(chaine))

    def marquerErreur(self, ligne, colonneOuEstLErreur, nbcellule):
        for colonne in colonneOuEstLErreur:
            for decalage in range(nbcellule):
                self.colorier([[ligne, colonne+decalage]], self.tableCentrale, self.rouge)


    def definirInterval(self):
       self.de= datetime.fromisoformat("2020-03-30")
       self.a= datetime.fromisoformat("2020-05-24")
       self.tableDate= self.dateDeA(self.de, self.a)
       self.tableDateJour= ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]*8
       self.tabHaute= [self.tableDate, self.tableDateJour]
        
    def dateDeA(self, de, a):
        listedate= [self.formatJJMMYYYY(str(de)[:10])]
        d= timedelta(days=1)
        while de<a:
           de= de+d
           listedate.append(self.formatJJMMYYYY(str(de)[:10])) 
        return listedate    

    def formatJJMMYYYY(self, chaine):
        chaine2= chaine.split("-")
        chaine2.reverse()
        return ".".join(chaine2)

    def minuteAHeureMinute(self, minutes):
       heure= str(minutes//60)
       minute= str(minutes-(60*int(heure)))
       if len(heure) == 1:
            heure= "0"+heure
       if len(minute) == 1:
            minute= "0"+minute
       return heure+":"+minute

    def compter(self, lettre, colonne):
        if lettre=="N" :
            self.N[colonne]=self.N[colonne]+1
        elif lettre=="R" :
            self.R[colonne]=self.R[colonne]+1
        elif lettre=="J" :
            self.J[colonne]=self.J[colonne]+1
        elif lettre=="I" :
            self.I[colonne]=self.I[colonne]+1
        elif lettre=="C" :
            self.C[colonne]=self.C[colonne]+1


    def modifierData(self, item):
        #modification des datas pour table central
        ligne= item.row()
        colonne= item.column()+self.debut
        texte= item.text()
        #enregistre la table 
        if self.tabCentrale[ligne][colonne] != texte: 
            self.tabCentrale[ligne][colonne]= texte
            self.initTabBas()
            self.heureTotal()
            self.chargerTable2(self.tabDroite, self.tableDroite)
            self.chargerTable(self.tabBas, self.tableBas)
            self.zoneCouleur()
            #checking des règles après changement
            self.regles(ligne, colonne)
            self.saveData() 

    def modifierDataTabGauche(self, item):
        ligne= item.row()
        colonne= item.column()
        texte= item.text()
        #enregistre la table 
        if self.tabGauche[ligne][colonne] != texte: 
            self.tabGauche[ligne][colonne]= texte
            self.initTabBas()
            self.heureTotal()
            self.chargerTable2(self.tabDroite, self.tableDroite)
            self.chargerTable(self.tabBas, self.tableBas)
            #checking des règles après changement
            self.regles(ligne, colonne)
            self.saveData() 

    def getJour(self, num):
        return self.planning[num-1]

    def getSemaine(self, num):
        num= num-1 #car on commence à partir de 0 avec les tableaux
        return self.planning[(num*7):(num+1)*7]

    def synchronisation(self):
        #première synchro table widget à table bas
        self.tableCentrale.horizontalScrollBar().valueChanged.connect(self.tableBas.horizontalScrollBar().setValue)
        self.tableBas.horizontalScrollBar().valueChanged.connect(self.tableCentrale.horizontalScrollBar().setValue)
        #deuxième synchro table widget à table droite
        self.tableCentrale.verticalScrollBar().valueChanged.connect(self.tableDroite.verticalScrollBar().setValue)
        self.tableDroite.verticalScrollBar().valueChanged.connect(self.tableCentrale.verticalScrollBar().setValue)
        #troisième synchro table widget à table gauche
        self.tableCentrale.verticalScrollBar().valueChanged.connect(self.tableGauche.verticalScrollBar().setValue)
        self.tableGauche.verticalScrollBar().valueChanged.connect(self.tableCentrale.verticalScrollBar().setValue)
        #première synchro table widget à table bas
        self.tableCentrale.horizontalScrollBar().valueChanged.connect(self.tableHaute.horizontalScrollBar().setValue)
        self.tableHaute.horizontalScrollBar().valueChanged.connect(self.tableCentrale.horizontalScrollBar().setValue)

    def comboLieu(self):
        print("changement de lieu")
        texte= str(self.combobox1.currentText())
        self.lieu= texte
        self.initTab()
        self.chargerToutesLesTab()

    def comboSemaine(self):
        print("changement de semaine")
        texte= str(self.combobox2.currentText())
        if texte != "Tous":
            tabTexte= texte.split(" ")
            nTabBas= []
            nTabCentrale= []
            nTabHaute= []
            n= int(tabTexte[1])
            self.debut= 7*(n-1)
            self.fin= (7*n)
            #redéfinition de la table droite
            nTabDroite=self.heureTotal2(self.tableDroite, self.debut, self.fin)
            #redéfinition de la table du haut
            nTabHaute.append(self.tabHaute[0][self.debut:self.fin]) 
            nTabHaute.append(self.tabHaute[1][self.debut:self.fin]) 
            #redéfinition de la table centrale
            for ligne in self.tabCentrale:
                nTabCentrale.append(ligne[self.debut:self.fin])
            #redéfinition de la table du bas
            for ligne in self.tabBas:
                nTabBas.append(ligne[self.debut:self.fin])
            #chargement des tables
            self.chargerTable2(nTabDroite, self.tableDroite)
            self.chargerTable(nTabCentrale, self.tableCentrale)
            self.chargerTable(nTabHaute, self.tableHaute)
            self.chargerTable(nTabBas, self.tableBas)
            self.marquerTable(self.tableCentrale)
            self.zoneCouleur2(nTabCentrale, nTabHaute)
        else:
            self.debut= 0
            self.fint= 56
            self.chargerToutesLesTab()

    def excel(self, fileName):
        workbook = xlsxwriter.Workbook(fileName+'.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.set_column(1, 60, 10)

        #chargement de la table de gauche
        for ligne in range(len(self.tabGauche)):
            if ligne%2 == 0:
                for colonne in range(3):
                    worksheet.write(ligne+1, colonne, self.tabGauche[ligne][colonne])            
                for colonne in range(len(self.tabCentrale[0])):
                    planning= self.tabCentrale[ligne][colonne]
                    couleur= self.dePlanningACouleur2(planning, workbook)
                    worksheet.write(ligne+1, colonne+3, planning, couleur)            
            else:
                for colonne in range(3):
                    worksheet.write(ligne+1, colonne, self.tabGauche[ligne][colonne], self.dePlanningACouleur2("C", workbook))            
                for colonne in range(len(self.tabCentrale[0])):
                    worksheet.write(ligne+1, colonne+3, self.tabCentrale[ligne][colonne], self.dePlanningACouleur2("C", workbook))            
                

        for colonne in range(len(self.tabHaute[0])):
                worksheet.write(0, colonne+3, self.deDateAJour(self.tabHaute[0][colonne]))            
        workbook.close()
        
    def deDateAJour(self, date):
        tab= date.split(".")
        return tab[0]

    def dePlanningACouleur(self, planning):
        couleur= self.blanc
        if planning == "N":
            couleur= self.bleu
        elif planning == "R":
            couleur= self.blanc
        elif planning == "J":
            couleur= self.jaune
        elif planning == "I":
            couleur= self.vert
        elif planning == "C":
            couleur= self.gris
        elif planning == " ":
            couleur= self.blanc
        else:
            couleur= self.rouge
        return couleur

    def dePlanningACouleur2(self, planning, workbook):
        #c'est comme dePlanningACouleur sauf que c'est pour la table excel exportée
        #on définit les couleurs ici car elles n'ont pas besoin d'être globales
        #en effet, on s'en sert que lorsqu'on exporte
        #définition des couleurs
        rouge = workbook.add_format({'bg_color': '#FF76A0', 'border':1})
        gris = workbook.add_format({'bg_color': '#B2BEC3', 'border':1})
        gris2 = workbook.add_format({'bg_color': '#DFE6E9', 'border':1})
        jaune = workbook.add_format({'bg_color': '#F6E58D', 'border':1})
        vert = workbook.add_format({'bg_color': '#7BED9F', 'border':1})
        bleu = workbook.add_format({'bg_color': '#74BAFF', 'border':1})
        blanc = workbook.add_format({'bg_color': '#FFFFFF',  'border':1})

        couleur= blanc
        if planning == "N":
            couleur= bleu
        elif planning == "R":
            couleur= blanc
        elif planning == "J":
            couleur= jaune
        elif planning == "I":
            couleur= vert
        elif planning == "C":
            couleur= gris2
        return couleur

    #va marquer les lignes d'un tableau selon le planning montré (N=bleu, J=jour, etc.)
    def marquerTable(self, table):
        tailleLigne= table.rowCount()
        tailleColonne= table.columnCount()
        for ligne in range(tailleLigne):
            for colonne in range(tailleColonne):
                planning= table.item(ligne, colonne).text()
                table.item(ligne, colonne).setBackground(self.dePlanningACouleur(planning))

    #va marquer la ligne d'un tableau selon le planning montré (N=bleu, J=jour, etc.)
    def marquerLigne(self, table, ligne):
        for colonne in range(table.columnCount()):
            planning= table.item(ligne, colonne).text()
            table.item(ligne, colonne).setBackground(self.dePlanningACouleur(planning))

    def fenetre(self, layoutContenu, largeur, longueur):
        fenetre = QDialog(self)
        fenetre.setLayout(layoutContenu)
        fenetre.setGeometry(100, 100, largeur, longueur)
        fenetre.setWindowTitle('window')
        fenetre.exec() 

    def deTroisColonneAUneColonne(self, tab):
        tabMilitaireUneColonne= []        
        for ligne in range(len(tab)):
            if ligne%2 == 0:
                tabMilitaireUneColonne.append(" ".join(tab[ligne]))
        return tabMilitaireUneColonne

    def getPlanningEtBloc(self, tab, ligne):
        doubleLigne= self.deLigneADoubleLigne(ligne)
        return [tab[doubleLigne[0]], tab[doubleLigne[1]]]

    def deLigneADoubleLigne(self, ligne):
        doubleLigne= [2*ligne, (2*ligne)+1]
        return doubleLigne

    def deDoubleLigneALigne(self, doubleLigne):
        #si c'est pair
        if doubleLigne%2 == 0:
            ligne= doubleLigne/2
            #si c'est impair
        else:
            ligne= (doubleLigne-1)/2

    #permet de créer rapidement un bouton sans se soucier de tout les détails
    def initButton(self, texte, callback):
        button = QPushButton(texte)
        button.clicked.connect(callback)
        return button
        
    def initOptionsButton(self):
        options= QGridLayout()
        buttonGererMilitaire= self.initButton("gérer militaire", self.gererMilitaire)
        buttonExporterExcel= self.initButton("exporter au format excel", self.saveFileDialog)
        buttonGererLieu= self.initButton("gérer lieu", self.gererLieu)
        options.addWidget(buttonGererMilitaire, 0, 0)
        options.addWidget(buttonExporterExcel, 1, 0)
        options.addWidget(buttonGererLieu, 2, 0)
        return options

    def gererMilitaire(self):
        #pour éviter tout bug si on click sur les boutons sans sélectionner
        self.ligne= 0
        #on défini tabGaucheUneColonne et tableGaucheUneColonne
        self.tabGaucheUneColonne= self.deTroisColonneAUneColonne(self.tabGauche)
        self.tableGaucheUneColonne= self.creerTable(400, 600)
        self.chargerTable2(self.tabGaucheUneColonne, self.tableGaucheUneColonne)
        self.tableGaucheUneColonne.cellClicked.connect(self.getCoordonnees)

        buttonPlus= self.initButton("+", self.gererMilitairePlus);
        buttonMoins= self.initButton("-", self.gererMilitaireMoins);

        buttonImporterListe= self.initButton("importer une liste", self.importerListe)

        haut= QGridLayout()
        haut.addWidget(buttonPlus, 0, 0)
        haut.addWidget(buttonMoins, 0, 1)

        layoutDroite= QGridLayout()
        layoutDroite.addLayout(haut, 0, 0)
        layoutDroite.addWidget(buttonImporterListe, 1, 0)

        layoutGererMilitaire= QGridLayout()
        layoutGererMilitaire.addLayout(layoutDroite, 0, 0)
        layoutGererMilitaire.addWidget(self.tableGaucheUneColonne, 1, 0)

        header = self.tableGaucheUneColonne.horizontalHeader()       
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        
        self.verouiller(self.selectColonne(0, self.tabGaucheUneColonne), self.tableGaucheUneColonne)
        
        self.fenetre(layoutGererMilitaire, 400, 700)
    
    def gererMilitairePlus(self):
        #index est là pour rajouter des espaces invisibles qui varient à la création d'un nouveau militaire
        #le but est de ne pas créer des infos redondantes qui risquent de mélanger la base de donnée
        index= " "*len(self.tabGauche)
        #creer les nouvelles entrées
        nouveauMilitaireTabGauche= [["grade", "nom", "prenom"+index], [" ", " ", " "]]
        nouveauMilitaireTabCentrale= [[" "]*56, [" "]*56]
        doubleLigne= self.deLigneADoubleLigne(self.ligne)
        
        #on ajoute le militaire aux tables à la ligne souhaitée
        self.tabGauche.insert(doubleLigne[0], nouveauMilitaireTabGauche[0])
        self.tabGauche.insert(doubleLigne[1], nouveauMilitaireTabGauche[1])

        self.tabCentrale.insert(doubleLigne[0], nouveauMilitaireTabCentrale[0])
        self.tabCentrale.insert(doubleLigne[1], nouveauMilitaireTabCentrale[1])

        #on ajoute le militaire à la base de donnée
        self.data.insert(str(self.ligne+1)+",'"+"','".join(nouveauMilitaireTabGauche[0])+"','"+self.cinquanteSixEspace+"', '"+self.cinquanteSixEspace+"'", self.lieu)

        #on ajuste les identifiants des soldat pour qu'ils apparaissent dans l'ordre
        self.data.resetID(self.tabGauche, self.lieu)#on commence à partir de la ligne suivante

        #on réarange la liste en une colonne des militaires
        self.tabGaucheUneColonne= self.deTroisColonneAUneColonne(self.tabGauche)
        self.chargerTable2(self.tabGaucheUneColonne, self.tableGaucheUneColonne)

        #on réarange la vue pour le reste des tables
        self.initTab()
        self.chargerToutesLesTab()

    def gererMilitaireMoins(self):
        doubleLigne= self.deLigneADoubleLigne(self.ligne)
        
        #on supprime le militaire aux tables à la ligne souhaitée
        tab= self.tabGauche.pop(doubleLigne[0])
        self.tabGauche.pop(doubleLigne[0])#on supprime au même endroit car la liste s'est réajustée toute seule

        #on supprime le militaire à la base de donnée
        self.data.remove(self.ligne, self.lieu)

        #on ajuste les identifiants des soldat pour qu'ils apparaissent dans l'ordre
        self.data.resetID(self.tabGauche, self.lieu)#on commence à partir de la ligne suivante

        #on réarange la liste en une colonne des militaires
        self.tabGaucheUneColonne= self.deTroisColonneAUneColonne(self.tabGauche)
        self.chargerTable2(self.tabGaucheUneColonne, self.tableGaucheUneColonne)

        #on réarange la vue pour le reste des tables
        self.initTab()
        self.chargerToutesLesTab()

    def getCoordonnees(self, ligne, colonne):
        self.ligne= ligne;
        self.colonne= colonne;

    def dePlanningALigne(self, planning):
        ligne=-1
        if planning == "N":
            ligne= 0
        elif planning == "R":
            ligne= 1
        elif planning == "J":
            ligne= 2
        elif planning == "I":
            ligne= 3
        elif planning == "C":
            ligne= 4
        return ligne

    def getColonneTabCentrale(self, colonne):
        tab= []
        for ligne in range(len(self.tabGauche)):
           tab.append(self.tabCentrale[ligne][colonne])
        return tab
            
    def getPlanningBloc(self, tab, ligne):
        doubleLigne= self.deLigneADoubleLigne(ligne)
        return [tab[doubleligne[0], tab[doubleLigne[1]]]]

    def journee(self, ligne, colonne):
        self.tabCentraleUneColonne= self.getColonneTabCentrale(colonne)
        self.tabPlanning= ["N", "R", "J", "I", "C"]
        self.tabBloc= []
        self.tabAdjacence= [[], [], [], [], []]
        for ligne in range(int(len(self.tabCentraleUneColonne)/2)):
            planningBloc= self.getPlanningEtBloc(self.tabCentraleUneColonne, ligne)
            ligneTabPlanning= self.dePlanningALigne(planningBloc[0])
            if ligneTabPlanning > -1:
                self.tabAdjacence[self.dePlanningALigne(planningBloc[0])].append(planningBloc[1])
                self.tabBloc.append(planningBloc[1])
        #on retir les éléments redondant de la table des Blocs
        self.tabBloc= list(set(self.tabBloc))
        #on print les résultats:
        #on définit la table croisée
        self.tabCroisee= [[" "]+self.tabBloc+["Total: "], ["N"], ["R"], ["J"], ["I"], ["C"], ["Total: "]]
        #on compte les occurences pour chaque élément de la table de bloc
        for bloc in self.tabBloc:
            for j in range(len(self.tabAdjacence)):
                   self.tabCroisee[j+1].append(self.tabAdjacence[j].count(bloc)) 
        #on fait le total des militaires par bloc
        for i in range(len(self.tabBloc)):
            self.tabCroisee[6].append(self.tabCroisee[1][i+1]+self.tabCroisee[2][i+1]+self.tabCroisee[3][i+1]+self.tabCroisee[4][i+1]+self.tabCroisee[5][i+1])
        #on fait le total des militaires par horaire
        print("tabCroisée: ", self.tabCroisee)
        for i in range(1, len(self.tabCroisee)):
            self.tabCroisee[i].append(sum(self.tabCroisee[i][1:]))
        self.tableCroisee= self.creerTable(600, 600)
        self.chargerTable(self.tabCroisee, self.tableCroisee)
        self.colorier(self.selectLigne(6, self.tabCroisee), self.tableCroisee, self.gris)
        self.colorier(self.selectColonne(len(self.tabBloc)+1, self.tabCroisee), self.tableCroisee, self.gris)
        self.layoutJournee= QGridLayout()
        self.layoutJournee.addWidget(self.tableCroisee)
        self.fenetre(self.layoutJournee, 600, 600) 
        
    def importerListe(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","Microsoft excel Files (*.xlsx)", options=options)
        if fileName:
            tab= self.nouvelleTabGauche(fileName)
            if len(tab) > 0:
                self.tabGauche= tab
                self.resetPlanningBloc()
                self.initTabBas()
                self.heureTotal()
                self.chargerTable(self.tabGauche, self.tableGauche)
                self.chargerTable(self.tabCentrale, self.tableCentrale)
                self.chargerTable2(self.tabDroite, self.tableDroite)
                self.chargerTable(self.tabBas, self.tableBas)
                self.zoneCouleur()
                #suppression des éléments du lieu
                self.data.viderToutLesMilitairesDUnLieu(self.lieu)
                for i in range(len(self.tabCentrale)):
                    #si on voit deux virgules, c'est normal
                    if i%2 == 0:
                        chaine= str(int(i/2))+",'"+("','".join(self.tabGauche[i]))+"','"+(",".join(self.tabCentrale[i]))+"','"+(",".join(self.tabCentrale[i+1]))+"'"
                        #print(chaine)
                        self.data.insert(chaine, self.lieu)
                

    def saveFileDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getSaveFileName(self,"QFileDialog.getSaveFileName()","","excel files (.xlsx);; All Files (*)", options=options)
        if fileName:
            self.excel(fileName)

    def nouvelleTabGauche(self, nom):
        xlsx_file = Path('SimData', nom)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active
        
        row_name = []
        for row in sheet.iter_rows(1, sheet.max_row):
                row_name.append([row[0].value, row[1].value, row[2].value])
                row_name.append([" ", " ", " "])
        if row_name[0][0] == "Grade" and row_name[0][1] == "Nom" and row_name[0][2] == "Prenom":
           return row_name
        else:
           return []

    def resetPlanningBloc(self):
        #remet à zéro les plannings et les blocs selons la table de gauche
        ligne= len(self.tabGauche)
        tab=[]
        for i in range(ligne):
           tab.append([" "]*56) 
        self.tabCentrale= tab

    def initOptions(self):
       #on obtient la liste des lieux existants
       self.data= Data()
       self.tabLieu= self.data.getLieu().split(",") 
       self.lieu= self.tabLieu[0]

    def plusDOption(self):
        options= self.initOptionsButton()
        #options= QGridLayout()
        self.fenetre(options, 200, 200)

    def gererLieu(self):
        print("gerer Lieu!")
        #pour éviter tout bug si on click sur les boutons sans sélectionner
        self.ligne= 0
        #on défini tabGaucheUneColonne et tableGaucheUneColonne
        self.tableLieu= self.creerTable(400, 600)
        self.chargerTable2(self.tabLieu, self.tableLieu)
        self.tableLieu.cellClicked.connect(self.getCoordonnees)

        buttonPlus= self.initButton("+", self.gererLieuPlus);
        buttonMoins= self.initButton("-", self.gererLieuMoins);

        layoutDroite= QGridLayout()
        layoutDroite.addWidget(buttonPlus, 0, 0)
        layoutDroite.addWidget(buttonMoins, 0, 1)

        layoutGererLieu= QGridLayout()
        layoutGererLieu.addLayout(layoutDroite, 0, 0)
        layoutGererLieu.addWidget(self.tableLieu, 1, 0)

        header = self.tableLieu.horizontalHeader()       
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        
        self.verouiller(self.selectColonne(0, self.tabLieu), self.tableLieu)
        
        self.fenetre(layoutGererLieu, 400, 700)
        
    def gererLieuPlus(self):
        print("Gerer Lieu plus")
        self.champPourLeNomDuNouveauLieu= QTextEdit() 
        valider= self.initButton("valider", self.gererLieuPlusValider)
        #annuler= self.initButton("annuler", self.gererLieuPlusAnnuler)
        bas= QGridLayout()
        bas.addWidget(valider, 0, 0)
        #bas.addWidget(annuler, 0, 1)
        box= QGridLayout()
        box.addWidget(self.champPourLeNomDuNouveauLieu, 0, 0)
        box.addLayout(bas, 1, 0)
        self.fenetre(box, 200, 200)


    def gererLieuPlusValider(self):
        #on crée le nouveau lieu
        print("Gerer lieu plus valider")
        nouveauNom= str(self.champPourLeNomDuNouveauLieu.toPlainText())
        self.tabLieu.append(nouveauNom)
        self.data.saveLieu(",".join(self.tabLieu))
        #on ajoute un militaire par défaut dans le nouveau lieu
        self.data.ajouterUnMilitaireParDefaut(nouveauNom)
        self.chargerTable2(self.tabLieu, self.tableLieu)
        self.combobox1.addItems([nouveauNom])

    def gererLieuMoins(self):
        print("Gerer Lieu Moins")
        self.tabLieu.pop(self.ligne)
        self.data.saveLieu(",".join(self.tabLieu))
        self.chargerTable2(self.tabLieu, self.tableLieu)
        self.combobox1.removeItem(self.ligne)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = App()
    sys.exit(app.exec_()) 


